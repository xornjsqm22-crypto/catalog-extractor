import streamlit as st
import ollama
import openpyxl
import base64
import fitz
import os

SAVE_PATH = 'Z:\\5조\\창호DB.xlsx'

HEADERS = [
    '기호', '회사', '창호명', '종류', '개폐방식',
    '프레임폭(mm)', '프레임두께(mm)', '벤트폭(mm)', '벤트두께(mm)',
    '유리종류', '유리두께(T)', '기밀등급', '차음성능(dB)', '방범성능', '방화여부', '비고'
]

PREFIX_MAP = {
    '창문': 'WD',
    '문': 'AD',
    '커튼월': 'CW',
}

def image_to_base64(image_path):
    with open(image_path, 'rb') as f:
        return base64.b64encode(f.read()).decode('utf-8')

def pdf_to_images(pdf_path, output_folder):
    doc = fitz.open(pdf_path)
    image_paths = []
    os.makedirs(output_folder, exist_ok=True)
    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=150)
        img_path = f"{output_folder}/page_{i+1}.jpg"
        pix.save(img_path)
        image_paths.append(img_path)
    return image_paths

def get_next_symbol(save_path, 종류):
    prefix = PREFIX_MAP.get(종류, 'ETC')
    try:
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
        count = sum(
            1 for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0] and str(row[0]).startswith(prefix)
        )
    except:
        count = 0
    return f"{prefix}-{str(count + 1).zfill(2)}"

def extract_spec(image_path):
    image_data = image_to_base64(image_path)
    prompt = """
    이 카탈로그 이미지에서 창호 제품 스펙을 추출해줘.
    스펙이 없는 페이지면 "스펙없음" 이라고만 답해.
    스펙이 있으면 아래 형식으로만 답해:

    회사:
    창호명:
    종류: (창문 / 문 / 커튼월 중 하나만)
    개폐방식:
    프레임폭(mm):
    프레임두께(mm):
    벤트폭(mm):
    벤트두께(mm):
    유리종류:
    유리두께(T):
    기밀등급:
    차음성능(dB):
    방범성능:
    방화여부: (Y 또는 N)
    비고:
    """
    response = ollama.chat(
        model='gemma3',
        messages=[{
            'role': 'user',
            'content': prompt,
            'images': [image_data]
        }]
    )
    return response['message']['content']

def save_to_excel(spec_text):
    if '스펙없음' in spec_text:
        return False

    data = {}
    for line in spec_text.strip().split('\n'):
        if ':' in line:
            key, _, value = line.partition(':')
            data[key.strip()] = value.strip()

    종류 = data.get('종류', '창문')
    기호 = get_next_symbol(SAVE_PATH, 종류)

    try:
        wb = openpyxl.load_workbook(SAVE_PATH)
        ws = wb.active
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    row = [
        기호,
        data.get('회사', ''),
        data.get('창호명', ''),
        종류,
        data.get('개폐방식', ''),
        data.get('프레임폭(mm)', ''),
        data.get('프레임두께(mm)', ''),
        data.get('벤트폭(mm)', ''),
        data.get('벤트두께(mm)', ''),
        data.get('유리종류', ''),
        data.get('유리두께(T)', ''),
        data.get('기밀등급', ''),
        data.get('차음성능(dB)', ''),
        data.get('방범성능', ''),
        data.get('방화여부', ''),
        data.get('비고', '')
    ]
    ws.append(row)
    wb.save(SAVE_PATH)
    return True

st.set_page_config(page_title="창호 카탈로그 자동 추출", page_icon="🏠")
st.title("🏠 창호 카탈로그 자동 추출")
st.write("PDF 또는 이미지를 올리면 스펙을 자동으로 추출해서 저장합니다.")
st.info(f"💾 저장 위치: {SAVE_PATH}")

uploaded_file = st.file_uploader(
    "카탈로그 파일 선택",
    type=['pdf', 'jpg', 'jpeg', 'png']
)

if uploaded_file:
    st.info(f"파일: {uploaded_file.name}")

    if st.button("🚀 자동 추출 시작"):
        temp_path = f"temp_{uploaded_file.name}"
        with open(temp_path, 'wb') as f:
            f.write(uploaded_file.read())

        saved_count = 0

        if uploaded_file.name.lower().endswith('.pdf'):
            with st.spinner('PDF 변환 중...'):
                images = pdf_to_images(temp_path, 'temp_pages')
            progress = st.progress(0)
            for i, img in enumerate(images):
                st.write(f'페이지 {i+1}/{len(images)} 분석 중...')
                spec = extract_spec(img)
                if save_to_excel(spec):
                    saved_count += 1
                progress.progress((i+1)/len(images))
        else:
            with st.spinner('이미지 분석 중...'):
                spec = extract_spec(temp_path)
                if save_to_excel(spec):
                    saved_count += 1

        st.success(f'✅ 완료! {saved_count}개 제품이 저장됐습니다!')

        with open(SAVE_PATH, 'rb') as f:
            st.download_button(
                '📥 창호DB.xlsx 다운로드',
                f,
                file_name='창호DB.xlsx'
            )
