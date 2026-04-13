import ollama
import openpyxl
import base64
import fitz  # pymupdf
import os
from pathlib import Path

def clean_number(value):
    if not value:
        return ''
    import re
    match = re.search(r'[\d.]+', str(value))
    return float(match.group()) if match else value

SAVE_PATH = '창호DB.xlsx'

HEADERS = [
    '기호', '회사', '창호명', '종류', '개폐방식',
    '프레임폭', '프레임두께', '벤트폭', '벤트두께',
    '유리종류', '유리두께', '기밀등급', '차음성능', '방범성능', '방화여부', '비고'
]

PREFIX_MAP = {
    '창문': 'WD',
    '문': 'AD',
    '커튼월': 'CW',
}

def image_to_base64(image_path):
    with open(image_path, 'rb') as f:
        return base64.b64encode(f.read()).decode('utf-8')

def pdf_to_images(pdf_path, output_folder):
    doc = fitz.open(pdf_path)
    image_paths = []
    os.makedirs(output_folder, exist_ok=True)
    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=150)
        img_path = f"{output_folder}/page_{i+1}.jpg"
        pix.save(img_path)
        image_paths.append(img_path)
    return image_paths

def get_next_symbol(save_path, 종류):
    prefix = PREFIX_MAP.get(종류, 'ETC')
    try:
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
        count = sum(
            1 for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0] and str(row[0]).startswith(prefix)
        )
    except:
        count = 0
    return f"{prefix}-{str(count + 1).zfill(2)}"

def extract_spec(image_path):
    image_data = image_to_base64(image_path)
    prompt = """
    이 카탈로그 이미지에서 창호 제품 스펙을 추출해줘.
    스펙이 없는 페이지면 "스펙없음" 이라고만 답해.
    스펙이 있으면 아래 형식으로만 답해:

    회사:
    창호명:
    종류: (창문 / 문 / 커튼월 중 하나만)
    개폐방식:
    프레임폭(mm):
    프레임두께(mm):
    벤트폭(mm):
    벤트두께(mm):
    유리종류:
    유리두께(T):
    기밀등급:
    차음성능(dB):
    방범성능:
    방화여부: (Y 또는 N)
    비고:
    """
    response = ollama.chat(
        model='gemma3',
        messages=[{
            'role': 'user',
            'content': prompt,
            'images': [image_data]
        }]
    )
    return response['message']['content']

def save_to_excel(spec_text, save_path=SAVE_PATH):
    if '스펙없음' in spec_text:
        print('  → 스펙 없는 페이지, 건너뜀')
        return False

    data = {}
    for line in spec_text.strip().split('\n'):
        if ':' in line:
            key, _, value = line.partition(':')
            data[key.strip()] = value.strip()

    종류 = data.get('종류', '창문')
    기호 = get_next_symbol(save_path, 종류)

    try:
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    row = [
        기호,
        data.get('회사', ''),
        data.get('창호명', ''),
        종류,
        data.get('개폐방식', ''),
        clean_number(data.get('프레임폭(mm)', '')),
        clean_number(data.get('프레임두께(mm)', '')),
        clean_number(data.get('벤트폭(mm)', '')),
        clean_number(data.get('벤트두께(mm)', '')),
        data.get('유리종류', ''),
        clean_number(data.get('유리두께(T)', '')),
        data.get('기밀등급', ''),
        clean_number(data.get('차음성능(dB)', '')),
        data.get('방범성능', ''),
        data.get('방화여부', ''),
        data.get('비고', '')
    ]
    ws.append(row)
    wb.save(save_path)
    print(f'  ✅ {기호} 저장완료')
    return True

def process_folder(folder_path):
    folder = Path(folder_path)
    files = list(folder.glob('*.pdf')) + \
            list(folder.glob('*.jpg')) + \
            list(folder.glob('*.jpeg')) + \
            list(folder.glob('*.png'))

    print(f'\n총 {len(files)}개 파일 처리 시작\n')

    for file in files:
        print(f'처리중: {file.name}')
        if file.suffix.lower() == '.pdf':
            temp_folder = f"temp_{file.stem}"
            images = pdf_to_images(str(file), temp_folder)
            for i, img in enumerate(images):
                print(f'  페이지 {i+1}/{len(images)} 분석중...')
                spec = extract_spec(img)
                save_to_excel(spec)
        else:
            spec = extract_spec(str(file))
            save_to_excel(spec)

    print('\n🎉 전체 처리 완료!')

if __name__ == '__main__':
    print('=== 창호 카탈로그 자동 추출 ===')
    print('1. 단일 파일 (PDF/이미지)')
    print('2. 폴더 전체 처리')
    choice = input('\n선택 (1 or 2): ')

    if choice == '1':
        path = input('파일 경로 입력: ')
        if path.endswith('.pdf'):
            images = pdf_to_images(path, 'temp_pdf')
            for i, img in enumerate(images):
                print(f'페이지 {i+1} 분석중...')
                spec = extract_spec(img)
                save_to_excel(spec)
        else:
            spec = extract_spec(path)
            save_to_excel(spec)
    elif choice == '2':
        folder = input('폴더 경로 입력: ')
        process_folder(folder)

    print(f'\n창호DB.xlsx 저장 완료!')
